import json
import pathlib
import re

import openpyxl


WORKBOOK = pathlib.Path("public/intranet-ahorro-plus/PROGRAMA AHORRO PLUSS.xlsm")
OUTPUT = pathlib.Path("app/intranet-ahorro-plus/data.ts")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def money(value):
    try:
        if value in (None, ""):
            return 0
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return 0


def maintenance_date(value):
    text = clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{7}", text):
        text = "0" + text
    if re.fullmatch(r"\d{8}", text):
        day, month, year = text[:2], text[2:4], text[4:]
        if 1 <= int(day) <= 31 and 1 <= int(month) <= 12:
            return f"{year}-{month}-{day}"
    return text


workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
base = workbook["BASE"]
clients = []

for row_number, row in enumerate(base.iter_rows(min_row=3, values_only=True), start=3):
    rut = clean(row[2] if len(row) > 2 else "")
    name = clean(row[3] if len(row) > 3 else "")
    if not rut or not name:
        continue

    maintenance_dates = [
        maintenance_date(row[index])
        for index in range(4, 9)
        if index < len(row) and maintenance_date(row[index])
    ]
    savings = [
        money(row[index])
        for index in range(9, 14)
        if index < len(row) and money(row[index])
    ]

    clients.append(
        {
            "id": f"cliente-{row_number}",
            "rut": rut,
            "name": name,
            "maintenanceDates": maintenance_dates,
            "maintenanceBranches": [],
            "savings": savings,
            "earned": money(row[14] if len(row) > 14 else 0),
            "redeemed": money(row[15] if len(row) > 15 else 0),
            "balance": money(row[16] if len(row) > 16 else 0),
            "agreement": clean(row[17] if len(row) > 17 else "") or "PLAZACARSERVICE",
            "phone": clean(row[18] if len(row) > 18 else ""),
            "emailStatus": clean(row[19] if len(row) > 19 else ""),
        }
    )

agreement_sheet = workbook["CONVENIOS"]
agreements = []
for row in agreement_sheet.iter_rows(values_only=True):
    agreement = clean(row[2] if len(row) > 2 else "")
    if agreement and agreement.upper() != "CAR SERVICE":
        agreements.append(agreement)

content = (
    "export type IntranetClient = { "
    "id: string; rut: string; name: string; maintenanceDates: string[]; "
    "maintenanceBranches?: string[]; savings: number[]; "
    "earned: number; redeemed: number; balance: number; "
    "agreement: string; phone: string; emailStatus: string; "
    "};\n\n"
    'export const sourceWorkbook = "/intranet-ahorro-plus/PROGRAMA AHORRO PLUSS.xlsm";\n\n'
    'export const branchViews = ["Luz Divina", "Alcazar de Torres", "El Tabo"] as const;\n\n'
    f"export const agreements = {json.dumps(agreements, ensure_ascii=False, indent=2)} as const;\n\n"
    f"export const initialClients: IntranetClient[] = {json.dumps(clients, ensure_ascii=False, indent=2)};\n"
)

OUTPUT.write_text(content, encoding="utf-8")
print(f"clients={len(clients)} agreements={len(agreements)}")
